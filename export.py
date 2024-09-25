import pandas as pd
from openpyxl import load_workbook 
from openpyxl.styles import PatternFill, Font, Color, Alignment  
from openpyxl import workbook 
from openpyxl.utils.dataframe import dataframe_to_rows
import random 
from rich.console import Console 
import os 
import emoji

console = Console() 

back_color = [
    "000000",
    "00008b",
    "006400",
    "8b0000",
    "a9a9a9",
    "800080",
    "000080",
    "008080",
    "800000",
    "36454f",
    ]

def exportToExcel(data,filename):
   
    # Convert the dictionary of dictionaries to a DataFrame
    
    df = pd.DataFrame.from_dict(data, orient='index')

    # Export the DataFrame to an Excel file
    document_folder = os.path.join(os.path.expanduser('~'),"Documents") 

    file_path = os.path.join(document_folder,filename) 
    df.to_excel(file_path, index=True,sheet_name="All")  # index=True includes the outer keys (Alice, Bob, etc.) as a column
    setStyleExcel(filename)
    # print("Data exported successfully to 'output.xlsx'")

    workbook = load_workbook(file_path) 
    worksheet = workbook.active 
    text_color = "FFFFFF"   # white 
    font = Font(color=text_color,name="Times New Roman",bold=True)  
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    worksheet.delete_cols(1) 

    # change the font of each rows 
    # make the column as wide as column text 

    for column in worksheet.columns:
        max_length = 0 
        column = list(column)
        for cell in column:
             try:
                 if len(str(cell.value)) > max_length:
                     max_length = len(str(cell.value)) 
             except:
                 pass 
             
        adjusted_width = max_length + 10
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        worksheet.row_dimensions[1].height = 30 

    header_fill = PatternFill(start_color="25383c", end_color="25383c", fill_type="solid")

    for cell in worksheet[1]:  # the first row contain the headers 
        cell.alignment = center_aligned_text 
        cell.fill = header_fill   # set headers background color
        cell.value = str(cell.value).upper()   # convert header text to uppercase 
        cell.font = font 
    worksheet.sheet_properties.tabColor = "36454F" 
    workbook.save(file_path)
    createSheets(data,filename) 
   
    # console.print(f"[bold red] If there is no information in the excel file, it means there is no agents witht give ids!![/bold red]")
    return file_path 

def createSheets(data,filename):
        document_folder = os.path.join(os.path.expanduser("~"),"Documents")
        file_path = os.path.join(document_folder,filename) 
        department_data = {} 
        text_color = "FFFFFF"   # white 
        center_aligned_text = Alignment(horizontal="center", vertical="center")
        header_font = Font(color=text_color,name="Times New Roman",bold=True)
        text_font =  Font(name='Times New Roman', size=12)

        workbook = load_workbook(file_path)
        
        for record_id, record in data.items():
            department = record['department'] 

            if department not in department_data:
                department_data[department] = [] 

            department_data[department].append(record) 
            
        for department, records in department_data.items():
            sheet = workbook.create_sheet(title=department) 
            df = pd.DataFrame(records)  
            for row in dataframe_to_rows(df,index=False, header=True): 
                sheet.append(row) 

            # Apply the font to all cells in the worksheet
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = text_font
                

            for column in sheet.columns:
                max_length = 0 
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value)) 
                    except:
                        pass 
                    
                adjusted_width = max_length + 8
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
                sheet.row_dimensions[1].height = 30 

                header_fill = PatternFill(start_color="25383c", end_color="25383c", fill_type="solid")

            for cell in sheet[1]:  # the first row contain the headers 
                cell.alignment = center_aligned_text 
                cell.fill = header_fill   # set headers background color
                cell.value = str(cell.value).upper()   # convert header text to uppercase 
                cell.font = header_font
            sheet.sheet_properties.tabColor = random.choice(back_color)
        # Save the updated workbook
        workbook.save(file_path) 

     
def setStyleExcel(filename):
  
    document_folder = os.path.join(os.path.expanduser("~"),"Documents")
    file_path = os.path.join(document_folder,filename) 
    # Load the workbook and select the active worksheet
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    # Define the font style (e.g., Times New Roman, size 12)
    font = Font(name='Times New Roman', size=12)
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    # Apply the font to all cells in the worksheet
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = font
    # Save the updated workbook
    workbook.save(file_path)



